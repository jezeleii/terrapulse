import os
from datetime import date
from pathlib import Path

import pandas as pd
import psycopg
from openpyxl import load_workbook
from psycopg.rows import dict_row
from dotenv import load_dotenv


def require_env(name: str) -> str:
    val = os.environ.get(name)
    if not val:
        raise ValueError(f"Missing required environment variable: {name}")
    return val


def resolve_as_of_date(value: float) -> date:
    try:
        year = int(value)
    except (TypeError, ValueError):
        return date.today()
    if year < 1900 or year > 2100:
        return date.today()
    return date(year, 1, 1)


def load_year_sections(xlsx_path: str) -> dict:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["PROJECTS"]
    row3 = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    row4 = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))

    labels = {
        "issued_vintage": "Credits issued by vintage year (when reduction/removals occurred). >> for credits issued by issuance year scroll to the very right >>",
        "retired": "Credits retired or cancelled in:",
        "remaining_vintage": "Credits remaining by vintage:",
        "issued_issuance": "Credits issued by issuance year (when the registry issued the credits). << for credits issued by vintage year, scroll left to the green columns <<",
    }

    label_indices = {}
    for i, val in enumerate(row3):
        for key, label in labels.items():
            if val == label:
                label_indices[key] = i

    missing = [key for key in labels if key not in label_indices]
    if missing:
        raise ValueError(f"Missing yearly section labels in workbook header: {missing}")

    sorted_starts = sorted((idx, key) for key, idx in label_indices.items())
    boundaries = {key: (idx, None) for key, idx in label_indices.items()}
    for pos, (idx, key) in enumerate(sorted_starts):
        next_idx = sorted_starts[pos + 1][0] if pos + 1 < len(sorted_starts) else len(row4)
        boundaries[key] = (idx, next_idx)

    sections = {}
    for key, (start_idx, end_idx) in boundaries.items():
        year_map = {}
        unknown_idx = None
        for i in range(start_idx, end_idx):
            value = row4[i]
            if isinstance(value, int):
                year_map[int(value)] = i
            elif isinstance(value, str) and "Unknown" in value:
                unknown_idx = i
        sections[key] = {"year_map": year_map, "unknown_index": unknown_idx}

    return sections


def coerce_numeric_series(series: pd.Series) -> pd.Series:
    return (
        series.replace({",": ""}, regex=True)
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0)
    )


def build_yearly_records(
    df: pd.DataFrame,
    project_ids: list[str],
    years: list[int],
    year_columns: dict,
    retired_unknown_idx: int | None,
) -> list[dict]:
    issued_vintage_df = df.iloc[:, year_columns["issued_vintage"]].apply(coerce_numeric_series)
    retired_df = df.iloc[:, year_columns["retired"]].apply(coerce_numeric_series)
    remaining_df = df.iloc[:, year_columns["remaining_vintage"]].apply(coerce_numeric_series)
    issuance_df = df.iloc[:, year_columns["issued_issuance"]].apply(coerce_numeric_series)

    retired_unknown = None
    if retired_unknown_idx is not None:
        retired_unknown = coerce_numeric_series(df.iloc[:, retired_unknown_idx])

    yearly_records: list[dict] = []
    for row_idx, project_id in enumerate(project_ids):
        for col_idx, year in enumerate(years):
            issued_vintage = float(issued_vintage_df.iat[row_idx, col_idx])
            retired_total = float(retired_df.iat[row_idx, col_idx])
            remaining_vintage = float(remaining_df.iat[row_idx, col_idx])
            issued_issuance = float(issuance_df.iat[row_idx, col_idx])

            if issued_vintage or retired_total or remaining_vintage or issued_issuance:
                yearly_records.append(
                    {
                        "project_id": project_id,
                        "year": year,
                        "issued_vintage": issued_vintage,
                        "retired_total": retired_total,
                        "remaining_vintage": remaining_vintage,
                        "issued_issuance": issued_issuance,
                    }
                )

        if retired_unknown is not None:
            unknown_value = float(retired_unknown.iat[row_idx])
            if unknown_value:
                yearly_records.append(
                    {
                        "project_id": project_id,
                        "year": -1,
                        "issued_vintage": 0,
                        "retired_total": unknown_value,
                        "remaining_vintage": 0,
                        "issued_issuance": 0,
                    }
                )

    return yearly_records


def main() -> None:
    # -------------------------------------------------------------------
    # 0) Setup
    # -------------------------------------------------------------------
    # Load env file located in the same folder as this script (etl/.env)
    base_dir = Path(__file__).resolve().parent
    load_dotenv(dotenv_path=base_dir / ".env")

    db_url = require_env("SUPABASE_DB_URL")

    # Default to etl/data/vcm_berkeley.xlsx unless overridden
    xlsx_path = os.environ.get(
        "BERKELEY_XLSX_PATH",
        str(base_dir / "data" / "vcm_berkeley.xlsx")
    )

    truncate_before_load = os.environ.get("TRUNCATE_BEFORE_LOAD", "0").strip() == "1"

    # -------------------------------------------------------------------
    # 1) Extract
    # -------------------------------------------------------------------
    df = pd.read_excel(
        xlsx_path,
        sheet_name="PROJECTS",
        header=3
    )

    # -------------------------------------------------------------------
    # 2) Transform
    # -------------------------------------------------------------------
    section_info = load_year_sections(xlsx_path)
    issued_vintage_years = sorted(section_info["issued_vintage"]["year_map"].keys())
    retired_years = sorted(section_info["retired"]["year_map"].keys())
    remaining_years = sorted(section_info["remaining_vintage"]["year_map"].keys())
    issuance_years = sorted(section_info["issued_issuance"]["year_map"].keys())

    if issued_vintage_years != retired_years or issued_vintage_years != remaining_years or issued_vintage_years != issuance_years:
        raise ValueError("Year columns mismatch across yearly sections; check the workbook layout.")

    year_columns = {
        "issued_vintage": [section_info["issued_vintage"]["year_map"][year] for year in issued_vintage_years],
        "retired": [section_info["retired"]["year_map"][year] for year in issued_vintage_years],
        "remaining_vintage": [section_info["remaining_vintage"]["year_map"][year] for year in issued_vintage_years],
        "issued_issuance": [section_info["issued_issuance"]["year_map"][year] for year in issued_vintage_years],
    }
    retired_unknown_idx = section_info["retired"]["unknown_index"]

    column_map = {
        "Project ID": "project_id",
        "Project Name": "project_name",
        "Voluntary Registry": "registry",
        "Voluntary Status": "status",
        "Scope": "scope",
        " Type": "project_type",  # source column literally has a leading space
        "Reduction / Removal": "reduction_removal",
        "Methodology / Protocol": "methodology",
        "Methodology Version": "methodology_version",
        "Region": "region",
        "Country": "country",
        "State": "state",
        "Project Developer": "developer",
        "Project Site Location": "operator",
        "Verifier": "verifier",
        "Registry Documents": "registry_documents",
        "Total Credits \nIssued": "issued_total",
        "Total Credits \nRetired": "retired_total",
        "Total Credits Remaining": "remaining_total",
        "Total Buffer \nPool Deposits": "buffer_deposits_total",
        "First Year of Project (Vintage)": "first_year_vintage",
    }

    missing = [c for c in column_map.keys() if c not in df.columns]
    if missing:
        cols_preview = "\n".join([str(c) for c in df.columns[:60]])
        raise ValueError(
            "Missing expected columns in PROJECTS sheet:\n"
            + "\n".join(missing)
            + "\n\nFirst 60 columns detected:\n"
            + cols_preview
            + "\n\nFix: update column_map keys to match the workbook headers."
        )

    project_df = df[list(column_map.keys())].rename(columns=column_map)

    num_cols = [
        "issued_total",
        "retired_total",
        "remaining_total",
        "buffer_deposits_total",
        "first_year_vintage",
    ]

    for col in num_cols:
        project_df[col] = (
            project_df[col]
            .replace({",": ""}, regex=True)
            .pipe(pd.to_numeric, errors="coerce")
            .fillna(0)
        )

    text_cols = [c for c in project_df.columns if c not in num_cols]
    for col in text_cols:
        project_df[col] = (
            project_df[col]
            .where(project_df[col].notna(), None)
            .astype(str)
            .str.strip()
            .replace({"nan": None, "None": None, "": None})
        )

    project_df["as_of_date"] = project_df["first_year_vintage"].apply(resolve_as_of_date)
    project_df = project_df.drop(columns=["first_year_vintage"])

    project_df = project_df[project_df["project_id"].notna()]
    project_df = project_df.drop_duplicates(subset=["project_id"])
    df = df.loc[project_df.index].reset_index(drop=True)
    project_df = project_df.reset_index(drop=True)

    print(f"[ETL] Parsed rows: {len(project_df)}")
    if "registry" in project_df.columns:
        print("[ETL] Top registries:")
        print(project_df["registry"].value_counts().head(10).to_string())

    yearly_records = build_yearly_records(
        df,
        project_df["project_id"].tolist(),
        issued_vintage_years,
        year_columns,
        retired_unknown_idx,
    )

    # -------------------------------------------------------------------
    # 3) Load
    # -------------------------------------------------------------------
    upsert_sql = """
    insert into public.vcm_projects (
      project_id, project_name, registry, status,
      scope, project_type, reduction_removal,
      methodology, methodology_version,
      region, country, state,
      developer, operator, verifier, registry_documents,
      issued_total, retired_total, remaining_total, buffer_deposits_total,
      as_of_date
    )
    values (
      %(project_id)s, %(project_name)s, %(registry)s, %(status)s,
      %(scope)s, %(project_type)s, %(reduction_removal)s,
      %(methodology)s, %(methodology_version)s,
      %(region)s, %(country)s, %(state)s,
      %(developer)s, %(operator)s, %(verifier)s, %(registry_documents)s,
      %(issued_total)s, %(retired_total)s, %(remaining_total)s, %(buffer_deposits_total)s,
      %(as_of_date)s
    )
    on conflict (project_id) do update set
      project_name = excluded.project_name,
      registry = excluded.registry,
      status = excluded.status,
      scope = excluded.scope,
      project_type = excluded.project_type,
      reduction_removal = excluded.reduction_removal,
      methodology = excluded.methodology,
      methodology_version = excluded.methodology_version,
      region = excluded.region,
      country = excluded.country,
      state = excluded.state,
      developer = excluded.developer,
      operator = excluded.operator,
      verifier = excluded.verifier,
      registry_documents = excluded.registry_documents,
      issued_total = excluded.issued_total,
      retired_total = excluded.retired_total,
      remaining_total = excluded.remaining_total,
      buffer_deposits_total = excluded.buffer_deposits_total,
      as_of_date = excluded.as_of_date;
    """

    records = project_df.to_dict("records")

    try:
        with psycopg.connect(db_url, row_factory=dict_row) as conn:
            with conn.cursor() as cur:
                if truncate_before_load:
                    print("[ETL] TRUNCATE_BEFORE_LOAD=1 → truncating vcm_projects")
                    cur.execute("truncate table public.vcm_projects;")
                    cur.execute("truncate table public.vcm_project_yearly;")
                    conn.commit()

                cur.executemany(upsert_sql, records)
                conn.commit()
                print(f"[ETL] Upserted {len(project_df)} rows into public.vcm_projects")

                if yearly_records:
                    delete_sql = "delete from public.vcm_project_yearly where project_id = any(%s);"
                    cur.execute(delete_sql, (project_df["project_id"].tolist(),))

                    yearly_upsert_sql = """
                    insert into public.vcm_project_yearly (
                      project_id, year, issued_vintage, retired_total, remaining_vintage, issued_issuance
                    )
                    values (
                      %(project_id)s, %(year)s, %(issued_vintage)s, %(retired_total)s, %(remaining_vintage)s, %(issued_issuance)s
                    )
                    on conflict (project_id, year) do update set
                      issued_vintage = excluded.issued_vintage,
                      retired_total = excluded.retired_total,
                      remaining_vintage = excluded.remaining_vintage,
                      issued_issuance = excluded.issued_issuance;
                    """
                    cur.executemany(yearly_upsert_sql, yearly_records)
                    conn.commit()
                    print(f"[ETL] Upserted {len(yearly_records)} rows into public.vcm_project_yearly")

                # -----------------------------------------------------------
                # 4) Refresh derived metrics
                # -----------------------------------------------------------
                cur.execute("select public.refresh_vcm_metrics();")
                conn.commit()
                print("[ETL] Refreshed materialised view: public.vcm_project_metrics")

    except Exception as e:
        raise RuntimeError(f"ETL failed: {e}") from e
    print("[ETL] Done.")


if __name__ == "__main__":
    main()
